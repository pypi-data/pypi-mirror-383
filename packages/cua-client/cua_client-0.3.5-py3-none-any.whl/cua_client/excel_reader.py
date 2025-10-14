import logging
from typing import Dict, List, Any, Union

import pandas as pd
from openpyxl import load_workbook
from pydantic import BaseModel

logger = logging.getLogger(__name__)


class ExcelReaderArgs(BaseModel):
    file_path: str
    sheet: Union[int, str, None] = None


class ExcelReaderFunction:
    def __init__(self, max_rows_per_sheet: int = 1000, max_chars: int = 100_000):
        self.max_rows_per_sheet = max_rows_per_sheet
        self.max_chars = max_chars

    # --------------------------
    # Helper functions
    # --------------------------
    def _inject_hyperlink_columns(self, df: pd.DataFrame, ws) -> pd.DataFrame:
        """Add extra columns with *_link suffix when a cell has a hyperlink."""
        if df.empty:
            return df

        for col_idx, col_name in enumerate(df.columns, start=1):
            link_values: List[str | None] = []
            has_any_link = False

            # Worksheet rows start at 2 (row 1 == header) => df index 0 refers to ws row 2
            for df_row_idx in range(len(df)):
                ws_row = df_row_idx + 2
                cell = ws.cell(row=ws_row, column=col_idx)
                if cell.hyperlink is not None:
                    link_values.append(cell.hyperlink.target)
                    has_any_link = True
                else:
                    link_values.append(None)

            if has_any_link:
                new_col_name = f"{col_name}_link"
                df[new_col_name] = link_values

        return df

    def _dataframe_to_text(self, df: pd.DataFrame, sheet_name: str) -> str:
        if df.empty:
            return f"Sheet: {sheet_name}\n(Empty)\n\n"

        limited = df.head(self.max_rows_per_sheet)
        csv_text = limited.to_csv(index=False)
        header = f"Sheet: {sheet_name} (rows shown: {len(limited)}/{len(df)})\n"
        return header + csv_text + "\n"

    # --------------------------
    # Public entry point
    # --------------------------
    def __call__(self, **kwargs) -> Dict[str, Any]:
        args = ExcelReaderArgs(**kwargs)
        path = args.file_path

        try:
            # Use read_only=False to ensure cells expose hyperlink metadata
            wb = load_workbook(path, read_only=False, data_only=True)

            # Determine sheets to process (file-agnostic, do not assume specific names)
            if args.sheet is None:
                sheet_names = wb.sheetnames
            else:
                if isinstance(args.sheet, str):
                    if args.sheet in wb.sheetnames:
                        sheet_names = [args.sheet]
                    else:
                        logger.warning(f"Requested sheet '{args.sheet}' not found. Falling back to all sheets.")
                        sheet_names = wb.sheetnames
                elif isinstance(args.sheet, int):
                    if 0 <= args.sheet < len(wb.sheetnames):
                        sheet_names = [wb.sheetnames[args.sheet]]
                    else:
                        logger.warning(f"Requested sheet index {args.sheet} out of range. Falling back to all sheets.")
                        sheet_names = wb.sheetnames
                else:
                    sheet_names = wb.sheetnames

            parts: List[str] = []

            for sname in sheet_names:
                ws = wb[sname]
                df = pd.read_excel(path, sheet_name=sname, engine="openpyxl")

                # Inject hyperlink columns (non-fatal if it fails)
                try:
                    df = self._inject_hyperlink_columns(df, ws)
                except Exception:
                    logger.warning("Hyperlink extraction failed; returning sheet without link columns", exc_info=True)

                parts.append(self._dataframe_to_text(df, str(sname)))

            text = "".join(parts)

            if len(text) > self.max_chars:
                text = text[: self.max_chars] + "\n... (truncated)"

            return {"text": text}

        except Exception as e:
            logger.exception("Failed to read Excel file")
            return {"error": str(e)}


