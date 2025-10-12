def coupon_csvdata_read(testdata_file_name: str,
                         force_index: int,
                         strain_index: int):
    """
    Read tensile coupon test data from a CSV file and extract force and strain columns.

    Args:
        testdata_file_name (str): Path to the CSV file containing the test data.
        force_index (int): 1-based column index for the force data.
        strain_index (int): 1-based column index for the strain data.

    Returns:
        tuple:
            df_new (pd.DataFrame): DataFrame containing only 'Force' and 'Strain' columns.
            sample_name (str): Name of the sample extracted from the file name (without extension).

    Notes:
        - The CSV file is expected to have a header in the first row.
        - The second row is skipped during reading (`skiprows=[1]`).
        - The first column is treated as the index column (`index_col=0`).
        - Force and strain columns are extracted based on the provided 1-based indices.
    """
    
    import pandas as pd
    df = pd.read_csv(testdata_file_name, skiprows=[1])
    Force = df[df.columns[force_index - 1]]
    Strain = df[df.columns[strain_index - 1]]
    df_new = pd.concat([Force, Strain], axis=1)
    df_new.columns = ['Force', 'Strain']
    sample_name = testdata_file_name[:-4]
    return sample_name, df_new

def coupon_test_analysis (sample_name: str,
                          Force: float,
                          Strain: float,
                          thickness: float = 2.5,
                          width: float = 10,
                          showfig: bool = True,
                          savefig: bool = False,
                          low_bound: float = 0.1,
                          up_bound: float = 0.3,
                 ):
    """
    Post-process a tensile coupon test and plot stress-strain curve.
     
    To active with Jupyter Lab, '%matplotlib widget' is required
    
    Args:
        Thickness (float): Specimen thickness in mm.
        Width (float): Specimen width in mm.
        file_name (str): CSV file containing test data.
        low_bound (float): Lower bound of elastic region as fraction of UTS.
        upper_bound (float): Upper bound of elastic region as fraction of UTS.

    Returns:
        fig (matplotlib.figure.Figure): Figure object containing the plot.
    """
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    #%matplotlib widget
    
    # Constants
    area = thickness * width  # Calculate the area of the specimen
    
    # Load tensile test data

    strain = Strain # Strain in mm/mm
    
    # Calculate stress and strain
    force = Force * 1000 # Convert kN to N
    stress = (force / area)  # N/mm^2 or MPa
    uts = stress.max()
    
    #find the data before uts
    idx_peak = np.argmax(stress)
    strain_up = Strain[:idx_peak+1]
    stress_up = stress[:idx_peak+1]
    
    #Boundary for [low bound] - [up bound] of uts
    lower_bound = low_bound * uts
    upper_bound = up_bound * uts
    
    elastic_reg = (lower_bound <= stress_up) & (stress_up <= upper_bound)
    
    stress_ela = stress_up[elastic_reg]
    strain_ela = strain_up[elastic_reg] 
    
    E, intercept = np.polyfit(strain_ela, stress_ela, 1)
    #print(f"Young's Modulus is: {E} MPa",)
    E_GPa = E / 1000  # Convert MPa to GPa
    #print(f"Intercept: {intercept} MPa")
   
    # Select over [lower bound] of UTS, as yield stress will over [lower bound] uts
    strain_new = strain
    stress_new = force / area
    mask = (lower_bound <= stress)
    strain_mask = strain_new[mask]
    stress_mask = stress_new[mask]
    
    offset_decimal = 0.002  # 0.2% in decimal
    offset_line = E * (strain_new - offset_decimal) + intercept

    #Find the Yield strength
    diff = stress_mask - offset_line
    cross_index = np.where(diff <= 0)[0][0] +1
    x1 = strain_mask[cross_index-1]
    x2 = strain_mask[cross_index]
    y1 = diff[cross_index-1]
    y2 = diff[cross_index]
    yield_strain = x1 - y1 * (x2 - x1) / (y2 - y1)
    yield_strength = np.interp(yield_strain, strain_new, stress_new)

    #Plot
    fig, ax = plt.subplots(figsize=(10,6))
    
    ax.plot(strain, stress, label='Stress-Strain Curve', color='blue')
    ax.plot(strain_new, offset_line, label='0.2% Offset Strain Line', color='yellow',linestyle = '-.')
    
    ax.axhline(y=yield_strength, label=f'Yield Strength = {yield_strength:.2f} MPa', color='green', linestyle = 'dotted')
    ax.axhline(y=uts, color='red', linestyle = '--', label=f'UTS = {uts:.2f} MPa')
    ax.plot(yield_strain, yield_strength, 'ro', label='Yield Point')
    
    ax.set_xlabel('Strain (mm/mm)')
    ax.set_ylabel('Stress (MPa)')
    ax.set_title(sample_name + ' Stress-Strain Curve with Mechanical Properties')
    ax.legend()
    ax.grid(True)
    ax.set_ylim(0, 1.2 *uts)

    #Show fig or not
    if showfig == True:
        plt.show()
    else:
        plt.close(fig)

    # Save fig or not
    if savefig == True:
        fig.savefig(sample_name, dpi=300, bbox_inches='tight')
    
    # Print results
    print(f"Sample: {sample_name}")
    print(f"Young's Modulus (E): {E:.2f} MPa")
    print(f"Ultimate Tensile Strength (UTS): {uts:.2f} MPa")
    print(f"Yield Strength: {yield_strength:.2f} MPa")
    print('-' * 40)
    
    # Prepare results dictionary
    results = {
        "E_MPa": E,
        "UTS_MPa": uts,
        "Yield_Strength_MPa": yield_strength
    }
    return sample_name, results

def coupon_sample_geodata_read (Excelfile_name: str):
    '''
    Reads sample geometric properties from an Excel file.

    Parameters
    ----------
    Excelfile_name : str
        The name of the Excel file to read, located in the current working directory.

    Returns
    -------
    list
        A list containing:
        - sample_file_name (sample name from cell A2 with ".csv" extension)
        - thickness (value from cell C2)
        - width (value from cell B2)
    '''
    
    from openpyxl import load_workbook
    from pathlib import Path
    sample_data_file = Path.cwd()/Excelfile_name
    wb = load_workbook(sample_data_file)
    ws = wb.worksheets[0]
    samples = []
    for row in ws.iter_rows(min_row=2, values_only=True):  
        if all(cell is None for cell in row):  # skip empty
            continue
        sample_file_name = str (row[0] + ".csv")
        sample_name = coupon_SampleDetails(row[0], row[1], row[2], sample_file_name)
        samples.append(sample_name)
    return samples

def coupon_batch_analysis(Coupon_geodata: str,
                          force_index: float,
                          strain_index:float,
                          showfig: bool = True,
                          savefig: bool = False):
    """
    Perform batch analysis on a list of tensile coupon samples and return their results.

    Parameters
    ----------
    Coupon_geodata : list of SampleDetails
        A list of SampleDetails objects, each containing:
        - sample_file_name : str
            Path to the sample CSV file.
        - thickness : float
            Sample thickness (mm).
        - width : float
            Sample width (mm).
    force_index : int
        Column index of the Force data in the CSV file (1-based).
    strain_index : int
        Column index of the Strain data in the CSV file (1-based).
    showfig : bool, optional
        Whether to display stress-strain plots during analysis. Default is True.
    savefig : bool, optional
        Whether to save the stress-strain plots to files. Default is False.

    Returns
    -------
    list of SampleAnalysisResults
        A list of SampleAnalysisResults objects, each containing:
        - sample_name : str
            Name of the processed sample.
        - modulus_of_elasticity : float
            Elastic modulus (E_GPa).
        - ultimate_tensile_strength : float
            Ultimate tensile strength (UTS_MPa).
        - yield_strength : float
            Yield strength (Yield_Strength_MPa).
    """
    SARS = []
    for Coupon_detail in Coupon_geodata:
        csvdata = coupon_csvdata_read(Coupon_detail.sample_file_name,force_index,strain_index)
        result = coupon_test_analysis(Coupon_detail.sample_file_name[:-4], 
                                      csvdata[1]['Force'],
                                      csvdata[1]['Strain'],
                                      Coupon_detail.thickness, 
                                      Coupon_detail.width, 
                                      showfig, 
                                      savefig)
        SAR = coupon_SampleAnalysisResults(result[0], result[1]['E_MPa'], result[1]['UTS_MPa'], result[1]['Yield_Strength_MPa'])
        SARS.append(SAR)
    return SARS

def coupon_results_save(Excelfile_name: str, analysis_results: list):
    """
    Save a list of sample analysis results into an Excel file, matching by sample name.

    Parameters
    ----------
    Excelfile_name: str
        Path to the Excel file to save the results.
    analysis_results : list
        A list of SampleAnalysisResults objects, each containing:
        - sample_name
        - modulus_of_elasticity
        - ultimate_tensile_strength
        - yield_Strength

    Notes
    -----
    - Assumes the Excel file has sample names in column A, starting from row 2.
    - Data will be written starting from column D (fourth column).
    - Overwrites the original Excel file.
    """
    from openpyxl import load_workbook

    wb = load_workbook(Excelfile_name)
    ws = wb.worksheets[0]

    # Build a dictionary for faster lookup
    result_dict = {res.sample_name: res for res in analysis_results}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column), start=2):
        excel_sample_name = row[0].value
        if excel_sample_name in result_dict:
            analysis_result = result_dict[excel_sample_name]
            values_to_write = [
                analysis_result.modulus_of_elasticity,
                analysis_result.ultimate_tensile_strength,
                analysis_result.yield_Strength
            ]
            for col_idx, value in enumerate(values_to_write, start=4):  # D列开始
                ws.cell(row=row_idx, column=col_idx, value=float(value))

    wb.save(Excelfile_name)
    print('The coupon test data analysis is complete.')

from dataclasses import dataclass
@dataclass
class coupon_SampleDetails:
    """
    Holds basic information for a sample.

    Attributes
    ----------
    sample_name : str
        Name or ID of the sample.
    width : float
        Width of the sample (geometric property, in consistent units).
    thickness : float
        Thickness of the sample (geometric property, in consistent units).
    sample_file_name : str
        Name of the associated data file for the sample, e.g., CSV file.
    """
    sample_name: str
    width: float
    thickness: float
    sample_file_name: str

@dataclass
class coupon_SampleAnalysisResults:
    """
    Stores the analysis results for a sample after mechanical testing.

    Attributes
    ----------
    sample_name : str
        Name or ID of the sample corresponding to the analysis.
    modulus_of_elasticity : float
        Elastic modulus of the sample (E, in GPa).
    ultimate_tensile_strength : float
        Maximum tensile strength of the sample (UTS, in MPa).
    yield_Strength : float
        Yield strength of the sample (in MPa).
    """
    sample_name: str
    modulus_of_elasticity: float
    ultimate_tensile_strength: float
    yield_Strength: float