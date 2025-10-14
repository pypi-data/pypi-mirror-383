"""
This module is used to export the data to CSV file.
With colorful and well-formatted csv file, we can easily analyze the data.
"""

import pandas as pd
import numpy as np
from typing import List, Any
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from ....utils.logger import ModernLogger


class Export2CSV(ModernLogger):
    def __init__(self, columns: List[str], data: List[List[Any]], export_path: str) -> None:
        super().__init__(name="Export2CSV")
        self.columns = columns
        self.data = data
        self.export_path = export_path

    def export2csv(self) -> None:
        """Export the data to CSV file with formatting
    
        Creates a well-formatted Excel file with:
        - Bold headers
        - Alternating row colors (white and light gray)
        - Auto-adjusted column widths
        - Proper alignment
        - Clean borders
        
        Args:
            columns: List[str] - Column headers
            data: List[List[Any]] - Data rows as list of lists
            export_path: str - Path to save the file
        
        Returns:
            None
        """
        # Create DataFrame
        data_list = list(self.data)
        df = pd.DataFrame(data_list, columns=self.columns)
        
        # Check if the path is valid
        if not self.export_path.endswith(('.xlsx', '.xls')):
            self.export_path = os.path.splitext(self.export_path)[0] + '.xlsx'
        
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Formatted Data"
        
        # Add data from DataFrame to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply formatting
                if r_idx == 1:  # Header row
                    # Bold font for headers
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:  # Data rows
                    # Alternating row colors (white and light gray)
                    if r_idx % 2 == 0:  # Even rows
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    # else odd rows remain white
                    
                    # Align text based on data type
                    if isinstance(value, (int, float, np.number)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(data_list) + 1, min_col=1, max_col=len(self.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Adjust column width (with some padding)
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50
        
        # Save the workbook
        wb.save(self.export_path)
        self.success(f"result successfully exported!")
        self.file_saved(self.export_path,"xlsx")


    def export2csv_html(self) -> None:
        """Export the data to HTML file with formatting
        
        Creates a well-formatted HTML file with:
        - Bold headers
        - Alternating row colors (white and light gray)
        - Responsive design
        
        Args:
            columns: List[str] - Column headers
            data: List[List[Any]] - Data rows as list of lists
            export_path: str - Path to save the file
        
        Returns:
            None
        """
        # Create DataFrame
        data_list = list(self.data)
        df = pd.DataFrame(data_list, columns=self.columns)
        
        # Generate HTML with styling
        html_string = """
        <html>
        <head>
            <style>
                table {
                    border-collapse: collapse;
                    width: 100%;
                    font-family: Arial, sans-serif;
                }
                th {
                    background-color: #D9D9D9;
                    color: black;
                    font-weight: bold;
                    text-align: center;
                    padding: 8px;
                    border: 1px solid #ddd;
                }
                td {
                    text-align: left;
                    padding: 8px;
                    border: 1px solid #ddd;
                }
                tr:nth-child(even) {
                    background-color: #F2F2F2;
                }
                tr:nth-child(odd) {
                    background-color: #FFFFFF;
                }
            </style>
        </head>
        <body>
        """
        
        # Convert DataFrame to HTML table with classes for styling
        html_string += df.to_html(index=False, classes='styled-table')
        html_string += """
        </body>
        </html>
        """
        
        # Ensure path has .html extension
        if not self.export_path.endswith('.html'):
            self.export_path = os.path.splitext(self.export_path)[0] + '.html'
        
        # Write HTML to file
        with open(self.export_path, 'w') as f:
            f.write(html_string)
        
        self.success(f"HTML file successfully exported to {self.export_path}")
        self.file_saved(self.export_path)
