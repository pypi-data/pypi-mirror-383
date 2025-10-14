from typing import Dict, List
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .logger import ModernLogger

class Report(ModernLogger):
    """
    A class for generating formatted Excel reports from pandas DataFrames.
    """
    
    def __init__(self, df: pd.DataFrame, output_file: str, metric_config: Dict[str, str], required_columns: List[str], color_scheme: Dict[str, Dict[str, str]], title: str="Report"):
        """Initialize the Report class."""
        super().__init__(name="Report")
        self.df = df
        self.output_file = output_file
        self.metric_config = metric_config
        self.required_columns = required_columns
        self.color_scheme = color_scheme
        self.title = title
        
    def create_report_excel(self):
        """
        Create a single Excel file with all datasets in one worksheet,
        with data_set as the leftmost column and merged for consecutive rows.
        
        Parameters:
        -----------
        df : pd.DataFrame
            Input dataframe containing all datasets
        output_file : str
            Output Excel file path
        metric_config : Dict[str, str]
            Dictionary mapping column names to their categories (e.g., 'local measure', 'global measure')
            Columns not specified will have no category (merged cells vertically)
        required_columns : List[str]
            List of columns that must appear first in the specified order
        color_scheme : Dict[str, Dict[str, str]]
            Color scheme for the report
            
        Returns:
        --------
        str
            Path to the generated Excel file
        """
        # Create a copy of the dataframe to avoid modifying the original
        # Sort by data_set to ensure consecutive rows belong to the same dataset
        report_df = self.df.copy().sort_values('data_set')
        
        # Ensure no duplicate columns
        report_df = report_df.loc[:, ~report_df.columns.duplicated()]
        
        # Organize metric_config by group to ensure columns in the same group are together
        grouped_metrics = defaultdict(list)
        for col, category in self.metric_config.items():
            if col in report_df.columns:  # Only add columns that exist in DataFrame
                grouped_metrics[category].append(col)
        
        # Ensure required_columns are at the front
        all_columns = self.required_columns.copy()
        
        # Add other columns by group order
        # Define group priority order
        group_order = ["compare", "global measure", "local measure"]
        
        # First add columns from groups with defined order
        for group in group_order:
            if group in grouped_metrics:
                all_columns.extend(grouped_metrics[group])
                # Remove added groups from grouped_metrics
                del grouped_metrics[group]
        
        # Add columns from remaining groups
        for group, cols in grouped_metrics.items():
            all_columns.extend(cols)
        
        # Add any remaining columns from DataFrame (not defined in metric_config)
        for col in report_df.columns:
            if col not in all_columns:
                all_columns.append(col)
        
        # Only keep columns that exist in DataFrame, ordered by all_columns
        final_cols = [col for col in all_columns if col in report_df.columns]
        report_df = report_df[final_cols]
        
        # Fill NA values with '—'
        report_df = report_df.fillna('—')
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.title
        
        # Define styles based on color scheme
        title_font = Font(size=14, bold=True, color=self.color_scheme['title']['font_color'])
        title_fill = PatternFill(
            start_color=self.color_scheme['title']['fill_color'], 
            end_color=self.color_scheme['title']['fill_color'], 
            fill_type="solid"
        )
        
        header_fill_level1 = PatternFill(
            start_color=self.color_scheme['header_level1']['fill_color'], 
            end_color=self.color_scheme['header_level1']['fill_color'], 
            fill_type="solid"
        )
        header_font_level1 = Font(bold=True, color=self.color_scheme['header_level1']['font_color'])
        
        header_fill_level2 = PatternFill(
            start_color=self.color_scheme['header_level2']['fill_color'], 
            end_color=self.color_scheme['header_level2']['fill_color'], 
            fill_type="solid"
        )
        header_font_level2 = Font(bold=True, color=self.color_scheme['header_level2']['font_color'])
        
        center_align = Alignment(horizontal='center', vertical='center')
        center_align_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write title
        ws['A1'] = self.title
        ws.merge_cells(f'A1:{get_column_letter(len(final_cols))}1')
        title_cell = ws['A1']
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = title_fill
        
        # Organize headers: determine which columns have categories and which don't
        categories = {}  # Dictionary to track which columns belong to which category
        no_category_cols = []  # List of columns without categories (to be merged vertically)
        
        for col in final_cols:
            if col in self.metric_config:
                category = self.metric_config[col]
                if category not in categories:
                    categories[category] = []
                categories[category].append(col)
            else:
                no_category_cols.append(col)
        
        # Prepare header rows
        # First row: categories
        category_headers = []
        for col in final_cols:
            if col in self.metric_config:
                category_headers.append(self.metric_config[col])
            else:
                category_headers.append('')  # No category
        
        # Second row: column names
        detail_headers = final_cols.copy()
        
        # Write first header row (categories)
        for col_idx, header in enumerate(category_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font_level1
            cell.alignment = center_align
            cell.fill = header_fill_level1
        
        # Merge cells for category headers
        for category, category_cols in categories.items():
            if len(category_cols) > 1:
                start_indices = [final_cols.index(col) + 1 for col in category_cols]
                start_indices.sort()  # Ensure indices are ordered
                
                # Check if these columns are consecutive
                if max(start_indices) - min(start_indices) + 1 == len(start_indices):
                    # Columns are consecutive, can merge
                    start_col = min(start_indices)
                    end_col = max(start_indices)
                    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
                else:
                    # Columns are not consecutive, need to handle each range separately
                    # Find consecutive ranges
                    ranges = []
                    current_range = [start_indices[0]]
                    
                    for i in range(1, len(start_indices)):
                        if start_indices[i] == start_indices[i-1] + 1:
                            current_range.append(start_indices[i])
                        else:
                            ranges.append(current_range)
                            current_range = [start_indices[i]]
                    
                    ranges.append(current_range)  # Add the last range
                    
                    # Merge each consecutive range
                    for r in ranges:
                        if len(r) > 1:
                            ws.merge_cells(start_row=2, start_column=min(r), end_row=2, end_column=max(r))
        
        # Write second header row (specific metrics)
        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font_level2
            cell.alignment = center_align
            cell.fill = header_fill_level2
        
        # Merge cells vertically for columns without categories
        for col in no_category_cols:
            col_idx = final_cols.index(col) + 1
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col  # Use the column name as the cell value
        
        # Find max values for each dataset and numeric metric to highlight later
        dataset_max_values = {}
        numeric_columns = []
        
        # Identify numeric columns by checking the first non-NA value in each column
        for col in final_cols:
            if col not in self.required_columns[:2]:  # Skip data_set and solution columns
                values = report_df[col].dropna()
                if len(values) > 0 and isinstance(values.iloc[0], (int, float)):
                    numeric_columns.append(col)
        
        # For each dataset, find the max value for each numeric metric
        for dataset in report_df['data_set'].unique():
            dataset_rows = report_df[report_df['data_set'] == dataset]
            dataset_max_values[dataset] = {}
            
            for metric in numeric_columns:
                # Filter out non-numeric values and find max
                metric_values = []
                for val in dataset_rows[metric]:
                    if isinstance(val, (int, float)) and val != '—':
                        metric_values.append(val)
                
                if metric_values:
                    dataset_max_values[dataset][metric] = max(metric_values)
                else:
                    dataset_max_values[dataset][metric] = None
        
        # Write data rows starting from row 4
        start_row = 4
        current_dataset = None
        dataset_start_row = start_row
        
        # Define row colors based on color scheme
        row_colors = self.color_scheme['row_colors']
        
        # Write each row and track when to merge dataset cells
        for row_idx, row in enumerate(report_df.itertuples(index=False), start_row):
            dataset = row[0]  # data_set is the first column
            
            # Check if dataset has changed
            if current_dataset != dataset:
                # If not the first row, merge cells for the previous dataset
                if current_dataset is not None:
                    if dataset_start_row < row_idx - 1:  # More than one row for the dataset
                        ws.merge_cells(start_row=dataset_start_row, start_column=1, 
                                    end_row=row_idx-1, end_column=1)
                        cell = ws.cell(row=dataset_start_row, column=1)
                        cell.alignment = center_align_wrap
                
                # Start tracking the new dataset
                current_dataset = dataset
                dataset_start_row = row_idx
            
            # Set background color based on dataset
            if dataset in row_colors:
                row_fill = PatternFill(start_color=row_colors[dataset], end_color=row_colors[dataset], fill_type="solid")
            else:
                row_fill = PatternFill(start_color=row_colors['default'], end_color=row_colors['default'], fill_type="solid")
            
            # Write all cells for the current row
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align
                
                # Apply row background color
                cell.fill = row_fill
                
                # Format numeric values with 2 decimal places (except 0 and 1)
                metric_col = final_cols[col_idx-1]
                if metric_col in numeric_columns and isinstance(value, (int, float)) and value not in (0, 1, '—'):
                    cell.number_format = '0.00'
        
        # Merge cells for the last dataset if needed
        if current_dataset is not None and dataset_start_row < row_idx:
            ws.merge_cells(start_row=dataset_start_row, start_column=1, 
                        end_row=row_idx, end_column=1)
            cell = ws.cell(row=dataset_start_row, column=1)
            cell.alignment = center_align_wrap
        
        # Get the last row with data
        max_row = len(report_df) + 3  # Title + 2 header rows + data rows
        
        # Auto-fit column widths based on content
        for col_idx in range(1, len(final_cols) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            
            # Find the maximum length in each column
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Set column width based on content (plus some padding)
            adjusted_width = max_length + 4
            ws.column_dimensions[column].width = adjusted_width
        
        # Apply font styling to dataset column values
        for row in range(start_row, max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True, color=self.color_scheme['dataset_highlight'])
        
        # Save the workbook
        wb.save(self.output_file)
        self.success(f"Report successfully generated!")
        self.file_saved(self.output_file,"xlsx")