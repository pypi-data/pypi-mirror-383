"""Excel file parser module

This module is responsible for parsing Excel files and converting them to JSON data
Supports complex nested data structures (lists, dictionaries, nested objects, etc.)
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import json
import os
from typing import Optional

from . import head_type
from .constants import HEAD_MARKER, DATA_MARKER, COMMENT_PREFIX
from .exceptions import (
    TableFileNotFoundError,
    InvalidFileFormatError,
    EmptyFileError,
    MissingHeaderMarkerError,
    MissingDataMarkerError,
)


class Head:
    """Header management class

    Responsible for parsing and managing the header structure of Excel tables
    Headers can have multiple rows and support nested column definitions

    Attributes:
        sheet: Excel worksheet object
        column: Number of columns
        filename: Filename (for error reporting)
        head: Root header object
        head_list: Header list, each column corresponds to a header object
    """
    def __init__(
        self, sheet: Worksheet, row: tuple[Cell, ...], filename: Optional[str] = None
    ) -> None:
        """Initialize header manager

        Args:
            sheet: Excel worksheet object
            row: Cell tuple of the first header row (excluding the #head marker in the first column)
            filename: Filename (optional, for error reporting)
        """
        self.sheet = sheet
        self.column = len(row)
        self.filename = filename
        # Create root header object (usually HeadObject type)
        self.head = head_type.head_creator(row[0], filename)
        # Initialize header list, each column points to the root header
        self.head_list: list[head_type.HeadType] = [self.head] * self.column

    def get_cell_max_col(self, cell: Cell) -> int:
        """Get the maximum column index of a cell

        If the cell is a merged cell, returns the last column index of the merged range
        Otherwise returns the cell's own column index

        Args:
            cell: The cell to check

        Returns:
            Column index (0-based, relative to the data starting column)
        """
        # Check if it's a merged cell
        for i in self.sheet.merged_cells.ranges:
            if cell.coordinate in i:
                return i.max_col - 1
        return cell.column - 1

    def row_parser(self, row: tuple[Cell, ...]) -> None:
        """Parse header row

        Handles multi-row headers, building nested header structures
        Merged cells represent a parent node covering multiple child columns

        Args:
            row: Cell tuple of the header row
        """
        i = 0
        while i < self.column:
            if row[i].value:
                h = head_type.head_creator(row[i], self.filename)
                j = self.get_cell_max_col(row[i])
                self.head_list[i].add_child(h)
                self.head_list[i:j] = [h] * (j - i)
                i = j
            else:
                i += 1


def get_data(filename: str) -> head_type.data:
    """Get parsed data from Excel file

    Parses an Excel file with structured headers and data, converting it into
    a nested Python data structure (dict/list/primitives).

    Args:
        filename: Path to the Excel file (.xlsx or .xlsm)

    Returns:
        Parsed data structure matching the header definitions

    Raises:
        ValueError: If filename is empty
        TableFileNotFoundError: If file doesn't exist
        InvalidFileFormatError: If file format is invalid or not .xlsx/.xlsm
        EmptyFileError: If file or worksheet is empty
        MissingHeaderMarkerError: If #head marker is missing or invalid
        MissingDataMarkerError: If #data marker is not found
    """
    # Validate input
    if not filename:
        raise ValueError("Filename cannot be empty")

    if not os.path.exists(filename):
        raise TableFileNotFoundError(filename)

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise InvalidFileFormatError(filename)

    # Load workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except Exception as e:
        raise InvalidFileFormatError(filename) from e

    if not workbook.active:
        raise EmptyFileError(filename)

    iter_rows = workbook.active.iter_rows()

    # Check first row
    try:
        first_row = next(iter_rows)
    except StopIteration:
        raise EmptyFileError(filename)

    # Validate #head marker
    first_cell = first_row[0]
    if first_cell.value != HEAD_MARKER:
        raise MissingHeaderMarkerError(
            first_cell, str(first_cell.value) if first_cell.value else "empty", filename
        )

    # Parse header
    head = Head(workbook.active, first_row[1:], filename)

    # Parse rows
    is_data = False
    data_root = None

    for row in iter_rows:
        if row[0].value == COMMENT_PREFIX:
            continue
        elif row[0].value == DATA_MARKER:
            is_data = True
            data_root = head.head.parse_data(row[1:], True, filename)
            continue

        if is_data:
            head.head.parse_data(row[1:], False, filename)
        else:
            head.row_parser(row[1:])

    # Validate data was found
    if data_root is None:
        raise MissingDataMarkerError(filename)

    return data_root


def get_json(filename: str, indent: int = 2) -> str:
    """Convert Excel file to JSON string

    Args:
        filename: Path to Excel file
        indent: JSON indentation level (default 2)

    Returns:
        Formatted JSON string
    """
    return json.dumps(get_data(filename), indent=indent)
