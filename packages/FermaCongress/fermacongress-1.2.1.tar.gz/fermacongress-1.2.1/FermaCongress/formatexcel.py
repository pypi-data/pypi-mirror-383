import os
import pandas as pd
from tqdm import tqdm
from dateutil import parser
from datetime import datetime, date, time
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def format(input_path: str = None, dataframe: pd.DataFrame = None, output_path: str = None, headers: bool = False, input_sheet: str = None, output_sheet: str = None):
    
    def clean_illegal_characters(value):
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value
    
    if input_path and dataframe is not None:
        raise ValueError("Provide either 'input_path' or 'dataframe', not both.")

    if output_path is None:
        raise ValueError("An 'output_path' must be specified.")

    if not str(output_path).lower().endswith('.xlsx'):
        raise ValueError("The 'output_path' must be an .xlsx file.")
    
    if input_path:
        file_extension = os.path.splitext(input_path)[1].lower()
        if file_extension == '.xlsx':
            wb = load_workbook(input_path)
            if input_sheet:
                if input_sheet not in wb.sheetnames:
                    raise ValueError(f"The sheet '{input_sheet}' does not exist in the input file.")
                ws = wb[input_sheet]
            else:
                ws = wb.active 
        elif file_extension == '.csv':
            df = pd.read_csv(input_path)
            wb = Workbook()
            ws = wb.active
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
        else:
            raise ValueError("The input file must be either .xlsx or .csv.")
        
    elif dataframe is not None:
        dataframe = dataframe.apply(lambda col: col.map(clean_illegal_characters))
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws.append(r)
            
    # Remove all other sheets (if input_path is an Excel file)
    if input_path and os.path.splitext(input_path)[1].lower() == '.xlsx':
        for sheet_name in wb.sheetnames:
            if sheet_name != ws.title:
                del wb[sheet_name]
    
    if output_sheet:
        ws.title = output_sheet  # Use the user-provided sheet name
    else:
        ws.title = "Sheet"  # Default to "Sheet" if no sheet_name is provided

    # Function to check if value is a date
    def is_date(value):
        return isinstance(value, (datetime, date, time))

    # Function to check if value is numeric
    def is_numeric(value):
        return isinstance(value, (int, float))
    
    # Retrieve the header row (once at the top) and convert column names to lowercase
    header_row = [cell.value.lower() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    
    # Total steps for progress bar
    total_steps = 9  # Adjust based on the number of main operations


    # Initialize the progress bar
    with tqdm(total=total_steps, desc="Formatting Progress") as pbar:
        
        # -------------------------------------------------------------------------------------------------
        # Step 1: Trimming whitespace
        # -------------------------------------------------------------------------------------------------
        try:
            trim_count = 0
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
                for j, cell in enumerate(row, start=1):
                    if isinstance(cell.value, str) and cell.value.strip():
                        original_value = cell.value
                        trimmed_value = original_value.strip()  # ONLY trims leading and trailing whitespace
                        if trimmed_value != original_value:
                            cell.value = trimmed_value
                            trim_count += 1
            pbar.update(1)
        except Exception as e:
            print(f"Error in Trimming Whitespace: {e}")



        # -------------------------------------------------------------------------------------------------
        # 2. Applying Formatting for the Data Rows
        """
        - Font Name: Manrope
        - Font Size: 10
        - Alignment: Horizontal - Left, Vertical - Top
        - Wrap Text - True
        - Row Height: 48.75
        
        """
        # -------------------------------------------------------------------------------------------------
        try:
            font = Font(name="Manrope", size=10)
            alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
                row_index = row[0].row
                ws.row_dimensions[row_index].height = 48.75
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
            pbar.update(1)
        except Exception as e:
            print(f"Error in formatting data rows: {e}")
            
        
        
        # -------------------------------------------------------------------------------------------------
        # 3. Applying Column Alignment Based on Column
        """
        - Center Alignment (Center, Center): Numeric Columns, 'priority'
        - Left Alignment (Left, Top): Text Columns
        - Right Alignment (Right, Bottom): Date Columns
        """
        # -------------------------------------------------------------------------------------------------

        try:
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            date_alignment = Alignment(horizontal="right", vertical="bottom", wrap_text=True)

            date_columns = {'start_date', 'start_time', 'end_date', 'end_time', 
                            'date of posting', 'date', 'start date', 'start time', 'end date', 'end time'}
            id_columns = {'display_id', 'display id', 'abstract_id', 'abstract id'}

            header_row = [cell.value.lower().strip() if isinstance(cell.value, str) else "" for cell in ws[1]]

            for col_index, col in enumerate(ws.iter_cols(min_row=2, max_col=ws.max_column), start=1):
                col_name = header_row[col_index - 1] if col_index - 1 < len(header_row) else ""

                if col_name in id_columns or col_name == 'priority' or all(is_numeric(cell.value) for cell in col if cell.value is not None):
                    alignment = center_alignment
                elif col_name in date_columns or all(is_date(cell.value) for cell in col if cell.value is not None):
                    alignment = date_alignment
                else:
                    alignment = left_alignment
                    
                for cell in col:
                    cell.alignment = alignment

            pbar.update(1)
        except Exception as e:
            print(f"Error in column alignment: {e}")
            
            
            
        # -------------------------------------------------------------------------------------------------
        # 4. Applying Column Widths
        """
        Default Width: 21
        Custom Widths:
        - Title Columns: 60
        - ID Columns: 12
        - Date Columns: 22
        """
        # -------------------------------------------------------------------------------------------------

        try:
            default_width = 21
            for col_index in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_index)
                ws.column_dimensions[col_letter].width = default_width

            title_columns = {'abstract', 'title', 'tweet text', 'abstract title', 'session title', 
                            'abstract_title', 'session_title', 'old abstract', 'old title', 'full abstract'}
            id_columns = {'internal_id', 'id', 'internal id', 'int id', 'session_id', 'session id', 
                        'display id', 'display_id'}
            date_columns = {'start_date', 'start_time', 'end_date', 'end_time', 'date of posting',
                            'date', 'start date', 'start time', 'end date', 'end time'}

            column_width_mapping = {**dict.fromkeys(title_columns, 60),
                                    **dict.fromkeys(id_columns, 12),
                                    **dict.fromkeys(date_columns, 22)}

            header_row = [cell.value.lower().strip() if isinstance(cell.value, str) else "" for cell in ws[1]]
            name_to_col_letter = {header_row[i]: get_column_letter(i + 1) for i in range(len(header_row)) if header_row[i]}
            for col_name, col_letter in name_to_col_letter.items():
                if col_name in column_width_mapping:
                    ws.column_dimensions[col_letter].width = column_width_mapping[col_name]
            pbar.update(1)
            
        except Exception as e:
            print(f"Error in applying column widths: {e}")
            
        
        
        # -------------------------------------------------------------------------------------------------
        # 5. Applying Conditional Formatting to Priority Column
        """
        - Very High: Light Blue (A5B3F7)
        - High: Light Green (39C7A5)
        - Medium: Yellow (FFCA42)
        - Low: Pink (EA4970)
        - Not Relevant: Light Grey (A2A2A7)
        - Internal: No Background Color
        """
        # -------------------------------------------------------------------------------------------------

        try:
            colors = {
                'Very High': 'A5B3F7',
                'High': '39C7A5',
                'Medium': 'FFCA42',
                'Low': 'EA4970',
                'Not Relevant': 'A2A2A7',
                'Internal': None
            }

            priority_order = ['Very High', 'High', 'Internal', 'Medium', 'Low', 'Not Relevant']

            if 'priority' in header_row:
                priority_col_index = header_row.index('priority') + 1
                priority_col_letter = get_column_letter(priority_col_index)

                for priority, color in colors.items():
                    if color:
                        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        rule = CellIsRule(operator='equal', formula=[f'"{priority}"'], stopIfTrue=True, fill=fill)
                        ws.conditional_formatting.add(f'{priority_col_letter}2:{priority_col_letter}{ws.max_row}', rule)

                # Sort Data Efficiently
                priority_sort_map = {priority: index for index, priority in enumerate(priority_order)}
                rows_to_sort = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

                rows_to_sort.sort(key=lambda row: priority_sort_map.get(row[priority_col_index - 1], float('inf')))

                for row_index, row_data in enumerate(rows_to_sort, start=2):
                    for col_index, value in enumerate(row_data, start=1):
                        ws.cell(row=row_index, column=col_index, value=value)

            pbar.update(1)

        except Exception as e:
            print(f"Error in conditional formatting for priority column: {e}")
            
        
        # -------------------------------------------------------------------------------------------------
        # 6. Applying Formatting in Hyperlinks
        """
        - Font Name: Manrope
        - Font Size: 10
        - Font Color: Blue (0000FF)
        - Underline: Single
        """
        # -------------------------------------------------------------------------------------------------

        try:
            hyperlink_font = Font(name="Manrope", size=10, color="0000FF", underline="single")
            hyperlink_columns = {}

            for column_name in ['abstract link', 'url', 'link', 'abstract_link', 'ferma link', 'tweet url']:
                if column_name in header_row:
                    hyperlink_columns[column_name] = header_row.index(column_name) + 1

            for col_index in hyperlink_columns.values():
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith(('http://', 'https://')):
                            cell.hyperlink = cell.value
                            cell.font = hyperlink_font

            pbar.update(1)

        except Exception as e:
            print(f"Error in hyperlink formatting: {e}")
            
        
        # -------------------------------------------------------------------------------------------------
        # 7. Applying Date and Time Formatting
        # -------------------------------------------------------------------------------------------------

        try:
            columns_formats = {
                'date': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'DDDD, DD MMMM YYYY'},
                'start_time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm AM/PM'},
                'end_time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm AM/PM'},
                'start time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm AM/PM'},
                'end time': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'hh:mm AM/PM'},  
                'start_date': {'initial': '%Y-%m-%d %H:%M:%S', 'desired': 'YYYY-MM-DD HH:MM:SS'},  
                'end_date': {'initial': '%Y-%m-%d %H:%M:%S', 'desired': 'YYYY-MM-DD HH:MM:SS'},  
                'date of posting': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'DD-MM-YYYY'},  
                'created_at': {'initial': "%Y-%m-%d %H:%M:%S", 'desired': 'DD-MM-YYYY'},  
            }

            unmatched_columns = set()

            for column_name, format_info in columns_formats.items():
                if column_name not in header_row:
                    continue
                column_index = header_row.index(column_name) + 1

                initial_format = format_info.get('initial')
                desired_format = format_info.get('desired')

                for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                    cell = row[0]
                    if cell.value:
                        try:
                            if not isinstance(cell.value, datetime):
                                try:
                                    dt = datetime.strptime(str(cell.value), initial_format) if initial_format else parser.parse(str(cell.value))
                                except ValueError:
                                    unmatched_columns.add(column_name)
                                    continue
                            else:
                                dt = cell.value
                            cell.value = dt
                            cell.number_format = desired_format

                        except Exception:
                            unmatched_columns.add(column_name)
            pbar.update(1)

        except Exception as e:
            print(f"Error in date formatting: {e}")
            
            
        # -------------------------------------------------------------------------------------------------
        # 8. Applying All Borders
        # -------------------------------------------------------------------------------------------------

        try:
            border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                row_border = [cell for cell in row]
                for cell in row_border:
                    cell.border = border
            pbar.update(1)
        except Exception as e:
            print(f"Error in applying borders: {e}")
        
        # -------------------------------------------------------------------------------------------------        
        # 9. Applying Formatting for the Header Row
        """
        - Font Name: Playfair Display Black
        - Font Size: 11
        - Font Color: White (FFFFFF)
        - Bold: True
        - Background Color: Blue (1E41EB)
        - Alignment: Horizontal - Center, Vertical - Center
        - Wrap Text - True
        - Row Height: 38
        """
        # -------------------------------------------------------------------------------------------------

        try:
            font = Font(name="Playfair Display Black", size=11, bold=True, color="FFFFFF")
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            fill = PatternFill(start_color="1E41EB", end_color="1E41EB", fill_type="solid")  

            for cell in ws[1]:
                cell.font = font
                cell.alignment = alignment
                cell.fill = fill

            ws.row_dimensions[1].height = 38
            
            if headers:
                header_row = [cell.value.title() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
                for idx, cell in enumerate(ws[1]):
                    cell.value = header_row[idx]

            pbar.update(1)
        except Exception as e:
            print(f"Error in header row formatting: {e}")
            
        # -------------------------------------------------------------------------------------------------
        
        
    wb.save(output_path)
    print()
    if unmatched_columns:
            print(f"• Unmatched Formats Detected: {list(unmatched_columns)}")
    pbar.close()
