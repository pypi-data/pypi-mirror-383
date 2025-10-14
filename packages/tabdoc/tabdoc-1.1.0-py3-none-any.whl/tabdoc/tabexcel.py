#!/usr/bin/env python3
# coding=utf-8

"""
@author: guoyanfeng
@software: PyCharm
@time: 19-2-11 下午6:14
"""
from collections import Counter
from io import BytesIO
from typing import MutableMapping, Sequence, Union

import tablib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import WriteOnlyCell

from path import Path

__all__ = ("ExcelWriter",)


class ExcelWriter(object):
    """
    excel book writer
    """

    def __init__(self, excel_name, excel_path=None, is_add_border: bool = False,
                 is_add_color: bool = False, is_add_alignment: bool = False, is_add_font: bool = False,
                 is_freeze_panes: bool = False):
        """
            excel book writer
        Args:
            excel_path: excel path
            excel_name: excel 名称
            is_add_border: 是否增加单元格边框，默认False
            is_add_color: 是否增加单元格颜色，默认False
            is_add_alignment: 是否增加单元格对齐方式，默认False
            is_add_font: 是否增加单元格字体，默认False
            is_freeze_panes: 是否冻结窗口，默认False
        """
        self.excel_path = excel_path
        self.excel_name = f"{excel_name}.xlsx"
        self.is_add_border: bool = is_add_border
        self.is_add_color: bool = is_add_color
        self.is_add_alignment: bool = is_add_alignment
        self.is_add_font: bool = is_add_font
        self.is_freeze_panes: bool = is_freeze_panes

        self.excel_book = tablib.Databook()
        self.merge_cells_index = {}
        self.sheet_names = Counter()  # 多个sheet name的映射，防止名称重复造成错误

    def __enter__(self):
        """

        Args:

        Returns:

        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """

        Args:

        Returns:

        """
        self.save()

    # noinspection DuplicatedCode
    @staticmethod
    def _reduce_datetimes(row):
        """Receives a row, converts datetimes to strings."""

        row = list(row)

        for i, val in enumerate(row):
            if hasattr(val, "strftime"):
                row[i] = val.strftime("%Y-%m-%d %H:%M:%S")
            elif hasattr(val, "isoformat"):
                row[i] = val.isoformat()
        return tuple(row)

    def add_sheet(self, sheet_name, sheet_data: list, merge_cells=None):
        """
        为excel添加工作表
        Args:
            sheet_name: 工作表的名称
            sheet_data: 工作表的数据， 必须是列表中嵌套元祖、列表或者字典（从records查询出来的数据库的数据）
            merge_cells: 要合并的单元格的索引, [(start_row, start_column, end_row, end_column)],最小值从1开始
        Returns:

        """
        sheet_data = sheet_data if sheet_data else [{}]
        #  处理sheet name可能出现重复的情况
        self.sheet_names[sheet_name] += 1
        sheet_name = sheet_name if self.sheet_names[sheet_name] == 1 else f"{sheet_name}{self.sheet_names[sheet_name]}"

        excel_sheet = tablib.Dataset(title=sheet_name)

        for row in sheet_data:
            if not isinstance(row, (MutableMapping, Sequence)):
                raise ValueError("sheet_data值数据类型错误,请检查")

        # 处理list或者tuple个别长度不一致的情况
        first = sheet_data[0]
        if isinstance(first, Sequence):
            for index, row in enumerate(sheet_data[1:], 1):
                diff = len(row) - len(first)
                if abs(diff) > 0:
                    if isinstance(row, list):
                        row.extend(["" for _ in range(diff)])
                    else:
                        sheet_data[index] = (*row, *["" for _ in range(diff)])

        if isinstance(first, MutableMapping):
            excel_sheet.headers = list(first.keys())
            for row in sheet_data:
                row = self._reduce_datetimes(row.values())
                excel_sheet.append(row)
        else:
            excel_sheet.headers = first
            for row in sheet_data[1:]:
                row = self._reduce_datetimes(row)
                excel_sheet.append(row)

        self.excel_book.add_sheet(excel_sheet)
        if merge_cells:
            verify_cells_index = []
            for val in merge_cells:
                verify_cells_index.extend(val)
            if min(verify_cells_index) < 1:
                raise ValueError("Min value is 1")
            self.merge_cells_index[sheet_name] = merge_cells

    # noinspection PyProtectedMember
    def export_book(self, ):
        """Returns XLSX representation of DataBook."""

        wb = Workbook(write_only=True)
        for i, dset in enumerate(self.excel_book._datasets):
            ws: Worksheet = wb.create_sheet()
            ws.title = dset.title if dset.title else 'Sheet%s' % i
            self.dset_sheet(dset, ws)
            # 合并单元格
            if ws.title in self.merge_cells_index:
                for ws_row_col in self.merge_cells_index[ws.title]:
                    ws.merge_cells(start_row=ws_row_col[0], start_column=ws_row_col[1], end_row=ws_row_col[2],
                                   end_column=ws_row_col[3])
                    if self.is_add_alignment:
                        ws._get_cell(ws_row_col[0], ws_row_col[1]).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True)
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    # noinspection PyProtectedMember
    def dset_sheet(self, dataset, ws: Worksheet):
        """Completes given worksheet from given Dataset."""
        _package = dataset._package(dicts=False)

        for i, sep in enumerate(dataset._separators):
            _offset = i
            _package.insert((sep[0] + _offset), (sep[1],))

        if self.is_freeze_panes:
            #  Export Freeze only after first Line
            ws.freeze_panes = 'A2'

        bold = Font(bold=True)
        for row_index, row_data in enumerate(_package, 1):
            row_cells = []
            for column_index, row_cell_value in enumerate(row_data, start=1):
                cell = WriteOnlyCell(ws)
                cell_horizontal, cell_vertical = None, None
                if isinstance(row_cell_value, dict):
                    cell_value: Union[str, int, float] = row_cell_value.get("value")
                    if self.is_add_color:
                        cell_color: str = row_cell_value.get("color", None)
                        if cell_color:
                            cell.fill = PatternFill("solid", fgColor=cell_color.lstrip("# "))
                    if self.is_add_alignment:
                        # 处理水平居中
                        cell_horizontal: str = row_cell_value.get("horizontal", None)
                        if cell_horizontal and cell_horizontal not in ("general", "left", "center", "right"):
                            cell_horizontal = "general"  # 默认对其方式
                        # 处理垂直居中
                        cell_vertical: str = row_cell_value.get("vertical", None)
                        if cell_vertical and cell_vertical not in ("top", "center", "bottom"):
                            cell_vertical = "center"  # 默认对其方式
                else:
                    cell_value = row_cell_value
                # 设置对齐方式
                if self.is_add_alignment:
                    cell.alignment = Alignment(wrap_text=True, horizontal=cell_horizontal, vertical=cell_vertical)
                # 增加边框单线，这里是固定的
                if self.is_add_border:
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # 处理None值
                cell.value = cell_value if cell_value is not None else ""
                # bold headers
                if self.is_add_font and row_index == 1 and dataset.headers:
                    cell.font = bold
                # bold separators
                elif self.is_add_font and len(row_data) < dataset.width:
                    cell.font = bold
                # 添加到行
                row_cells.append(cell)
            else:
                ws.append(row_cells)

    def save(self, ):
        """
        保存工作簿
        Args:
        Returns:

        """
        if self.excel_path is None:
            file_path = self.excel_name
        else:
            file_path = Path(self.excel_path).joinpath(self.excel_name).abspath()

        with open(file_path, "wb") as f:
            f.write(self.export_book())
