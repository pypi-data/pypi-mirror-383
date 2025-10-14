# -*- coding: utf-8 -*-
# @Author: maoyongfan
# @email: maoyongfan@163.com
# @Date: 2025/7/8 09:56
import zipfile
import os
import shutil
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook

INPUT_XLSX = "/Users/maoyongfan/Downloads/xlsx_bomb_based_on_input.xlsx"
OUTPUT_XLSX = "/Users/maoyongfan/Downloads/xlsx_bomb_based_on_input_2.xlsx"
TEMP_DIR = "temp_xlsx"
REPEAT_ROW = 200000  # 重复行数（越大炸弹威力越强）

def generate_xlsx_by_copying_last_row():
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    # 获取最后一行的数据内容
    last_row_idx = ws.max_row
    last_row_values = [cell.value for cell in ws[last_row_idx]]

    # 新建工作簿
    new_wb = Workbook()
    new_ws = new_wb.active

    for i in range(1, REPEAT_ROW + 1):
        new_ws.append(last_row_values)
        if i % 10000 == 0:
            print(f"已写入 {i} 行...")

    new_wb.save(OUTPUT_XLSX)
    print(f"[+] 已生成包含 {REPEAT_ROW} 行的新表格: {OUTPUT_XLSX}")

if __name__ == "__main__":
    generate_xlsx_by_copying_last_row()